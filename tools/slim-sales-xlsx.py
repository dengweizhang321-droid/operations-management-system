import os
from openpyxl import load_workbook, Workbook


def main() -> None:
    src = os.environ["IN_XLSX"]
    dst = os.environ["OUT_XLSX"]
    required = [
        "网店订单号",
        "发货仓库",
        "销售渠道",
        "货品编号",
        "货品名称",
        "数量",
        "单价",
        "下单时间",
        "货品成本",
        "分摊后单价",
        "分摊后金额",
        "费用分摊",
        "毛利",
        "毛利率",
        "未税毛利",
        "未税毛利率(%)",
        "发货时间",
    ]

    print({"stage": "load", "src": src}, flush=True)
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True)

    header = None
    header_row = 0
    for idx, row in enumerate(rows, start=1):
        values = [str(value).strip() if value is not None else "" for value in row]
        if all(name in values for name in ["发货仓库", "销售渠道", "货品编号", "货品名称", "数量"]):
            header = values
            header_row = idx
            break
    if header is None:
        raise RuntimeError("header not found")

    index = {name: header.index(name) for name in required if name in header}
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError("missing columns: " + ",".join(missing))

    out = Workbook(write_only=True)
    output_sheet = out.create_sheet(title="sheetTitle")
    output_sheet.append(required)

    count = 0
    for row in rows:
        output_row = [row[index[name]] if index[name] < len(row) else None for name in required]
        if any(value is not None and value != "" for value in output_row):
            output_sheet.append(output_row)
            count += 1
            if count % 10000 == 0:
                print({"stage": "copy_rows", "rows": count}, flush=True)

    out.save(dst)
    wb.close()
    print({"stage": "done", "headerRow": header_row, "rows": count, "output": dst}, flush=True)


if __name__ == "__main__":
    main()
